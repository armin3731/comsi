# imports =============================
import pandas as pd #Managing data tables, columns and rows
import numpy as np #library for numerical matrix and operators
# import matplotlib.pyplot as plt #Showing graphs and plots
import os #Paths and operating system addresses
# import openpyxl #library for working with excel file
# from openpyxl import load_workbook #loading Excel File
# from openpyxl.utils import get_column_letter #Reading values of Excel data cells
import networkx as nx #Creating netwrok graphs
# from matplotlib import cm #Color maps in graphs
from scipy.stats import spearmanr,pearsonr
from scipy import stats
# import time
import math
from openpyxl import load_workbook, Workbook
import openpyxl
import json
# import pyreadr

# #SPARCC imports
# from analysis_methods import basis_corr
# from io_methods import read_txt, write_txt




# Functions ==================================

def unique(list1):
    '''
    a function to get unique values

    INPUT:
        list1 : a list to find unique values from

    OUTPUT:
        unique_vals : a list of unique values
    '''
    unique_vals = []
    # insert the list to the set
    list_set = set(list1)
    # convert the set to the list
    unique_list = (list(list_set))
    for x in unique_list:
        unique_vals.append(x)
    return unique_vals


def Signaling_Metabolite_v9(job_id, dataframe_abundance, dataframe_metabolite, dataframe_metadata, filter_params, comp_fact,comp_fact_comb,Sample_number=50,Iteration_number=15):
    '''
    Finds Signaling Metabolites using the chosen conditions and databases.

    Input(s):
    dataframe_abundance : Dataframe of Abundance database
    dataframe_metabolite : Dataframe of UPEX database
    dataframe_metadata : Dataframe of META-Data database
    filter_params : Filtering Parametes. this will filter samples as prefered
    comp_fact : the column in META-Data that comparison in taking on
    comp_fac_comb : Combined Comparison Factor. two by two combined of comparison factors.
                    ex. if comparison factor is chosen to difrentiate samples by their
                    Country and there are samples form more than two Countries in our database
                    it will compare signaling metabolite for each two Country
    Sample_number : number of sample in each iteration for Multi-Dimension comparison.
                    ##TODO: if =1 it will use One-Dimension comparison.
    Iteration_number : number of iterations for calculation Z-Score

    Output(s):
         ---         : ---
    '''

    #Define the Graph===============
    G_metabolic=nx.Graph(name='Metabolic_Graph')

    #Create JSON file for Network Presentation
    JSON_net_2node = {'nodes':[],'links':[]}

    #Add Nodes======================
    # there are two types of nodes: 1-Metabolites(Red Nodes) 2-Species(Blue Nodes)

    #Species NODES:
    species_nodes = dataframe_metabolite.columns.to_list()
    print('Spices Nodes = ',len(species_nodes))
    G_metabolic.add_nodes_from([(x,{'type':'Spc','color':'blue'}) for x in species_nodes])
    for each_species_node in species_nodes:
        JSON_net_2node['nodes'].append({"name":"%s"%(each_species_node), "size":"%f"%(1.0), "color":"#00c"})
    # if len(dataframe_abundance1.columns.to_list())==1:
    #Metabolite NODES:
    metabolite_nodes = dataframe_metabolite.index.to_list()
    print('Metabolite Nodes = ',len(metabolite_nodes))
    G_metabolic.add_nodes_from([(x,{'type':'Meta','color':'red'}) for x in metabolite_nodes])
    for each_metabolite_nodes in metabolite_nodes:
        JSON_net_2node['nodes'].append({"name":"%s"%(each_metabolite_nodes), "size":"%f"%(1.0), "color":"#c00"})

    #Add Edges======================
    # if len(dataframe_abundance1.columns.to_list())==1:
    for each_species in dataframe_metabolite.columns.to_list():
        dataframe_metabolite_filter = dataframe_metabolite[dataframe_metabolite[each_species] != 0]
        for each_metabolite in dataframe_metabolite_filter.index.to_list():
            G_metabolic.add_edge(each_species,each_metabolite)
            JSON_net_2node['links'].append({"source":"%s"%(each_metabolite),"target":"%s"%(each_species)})

    #Save JSON_net_2node
    path_json = os.path.join('Output_folder','Network_Normal','%s.json'%(job_id))
    with open(path_json, 'w') as fp:
        json.dump(JSON_net_2node, fp)
    # print(someerroor)
    # else:
    #     for i_species in range(len(dataframe_metabolite.columns.to_list())):
    #         for j_species in range(i_species,len(dataframe_metabolite.columns.to_list())):
    #             if i_species==j_species:
    #                 continue

    #             i_sepcies_name = dataframe_metabolite.columns.to_list()[i_species]
    #             j_sepcies_name = dataframe_metabolite.columns.to_list()[j_species]

    #             dataframe_metabolite_filter = dataframe_metabolite[(dataframe_metabolite[i_sepcies_name] != 0) & (dataframe_metabolite[j_sepcies_name] != 0)]
    #             if len(dataframe_metabolite_filter.index.to_list())!=0:
    #                 G_metabolic.add_edge(i_sepcies_name,j_sepcies_name, metabolite_name=)


    #Graph Info=====================
    print(nx.info(G_metabolic))
    for each_comp_fact in comp_fact_comb:

    #     try:
        print(each_comp_fact,'*-*-*-*-*-*-*-*-*--*-*')
        Metabolite_Score_eachpopulation = []
        dataframe_abundance1,dataframe_abundance2, dataframe_metabolite_corrected = diff_data_creator_v6_OneDim(dataframe_abundance, dataframe_metabolite, dataframe_metadata,
                                                                                                            filter_params=filter_params, comp_fact=comp_fact ,comp_fact_value=each_comp_fact)

        Metabolite_Score_eachpopulation, _ = iterative_zscore(dataframe_abundance1,dataframe_abundance2,dataframe_metabolite_corrected,G_metabolic,sample_number=Sample_number,iteration=Iteration_number)
        if Metabolite_Score_eachpopulation==None:
            continue
        Metabolite_Score_Corrected_eachpopulation = Zscore_metabolite_correcter(Metabolite_Score_eachpopulation)
        Metabolite_dataframe_values = [[list(Metabolite_Score_Corrected_eachpopulation.values())[x],list(Metabolite_Score_eachpopulation.values())[x]] for x in range(len(Metabolite_Score_eachpopulation))]
        dataframe_Z_Scores = pd.DataFrame( Metabolite_dataframe_values , columns=['Z_metabolite_Corrected','Z_metabolite'],index=list(Metabolite_Score_eachpopulation.keys()))
        dataframe_Z_Scores.to_csv('app_files/temp/%s_%s_%s.csv'%(str(job_id),str(each_comp_fact[0].split('/')[0]),str(each_comp_fact[1].split('/')[0])))





def diff_data_creator_v6_OneDim(dataframe_abundance, dataframe_metabolite, dataframe_metadata,filter_params, comp_fact ,comp_fact_value):

    print('Samples in df_Abundance1 : ', len(dataframe_abundance.columns))
    dataframe_abundance1 = filter_data_aundance_v6_OneDim(dataframe_abundance.copy(), dataframe_metadata.copy() ,
                                                               filter_params, comp_fact ,comp_fact_value[0])
    print('Samples in df_Abundance2 : ', len(dataframe_abundance.columns))
    dataframe_abundance2 = filter_data_aundance_v6_OneDim(dataframe_abundance.copy(), dataframe_metadata.copy() ,
                                                               filter_params, comp_fact ,comp_fact_value[1])
    print('total Metabolites : ', len(dataframe_metabolite.index))

    #Remove Not Connected Metabolites and Species
    # find Metabolites that are not connected to any species and
    #      are all zero for all of species (not even a single +1
    #      or -1 for one specie). These Metabolites are called
    #      Redundant_Metabolites
    Redundant_Metabolites =(dataframe_metabolite==0).all(axis=1)[(dataframe_metabolite==0).all(axis=1)==True].index.to_list()
    # Remove redundant metabolites (Not Connected Metabolites)
    dataframe_metabolite = dataframe_metabolite.drop(Redundant_Metabolites)

    # find Redundant Species as well (Not Connected Species)
    dataframe_metabolite = dataframe_metabolite.T
    Redundant_Spices =(dataframe_metabolite==0).all(axis=1)[(dataframe_metabolite==0).all(axis=1)==True].index.to_list()
    dataframe_metabolite =  dataframe_metabolite.drop(Redundant_Spices)
    dataframe_metabolite = dataframe_metabolite.T


    #We only need MSPs which are shared with Metabolite data
    metabolite_MSPs = dataframe_metabolite.columns.to_list()
    abundance_MSPs = dataframe_abundance.index.to_list()
    if len(abundance_MSPs) > len(metabolite_MSPs):
        Redundant_MSPs = set(abundance_MSPs) - set(metabolite_MSPs)
        dataframe_abundance1 = dataframe_abundance1.drop(Redundant_MSPs)
        dataframe_abundance2 = dataframe_abundance2.drop(Redundant_MSPs)
    elif len(metabolite_MSPs) > len(abundance_MSPs) :
        Redundant_MSPs = set(metabolite_MSPs) - set(abundance_MSPs)
        dataframe_metabolite = dataframe_metabolite.T
        dataframe_metabolite = dataframe_metabolite.drop(Redundant_MSPs)
        dataframe_metabolite = dataframe_metabolite.T

    #We only need MSPs which are shared with Abundace1 and Abundance2
    abundance1_MSPs = dataframe_abundance1.index.to_list()
    abundance2_MSPs = dataframe_abundance2.index.to_list()
    if len(abundance1_MSPs) > len(abundance2_MSPs):
        Redundant_MSPs = set(abundance1_MSPs) - set(abundance2_MSPs)
        dataframe_abundance1 = dataframe_abundance1.drop(Redundant_MSPs)
    elif len(abundance2_MSPs) > len(abundance1_MSPs) :
        Redundant_MSPs = set(abundance2_MSPs) - set(abundance1_MSPs)
        dataframe_abundance2 = dataframe_abundance2.drop(Redundant_MSPs)
    # print('dataframe_abundance1',dataframe_abundance1)
    # print('number of shared species',len(dataframe_abundance2.index.to_list()),len(dataframe_metabolite.columns.to_list()))
    # print('number of shared Metabolites',len(dataframe_metabolite.index.to_list()))
    return dataframe_abundance1,dataframe_abundance2,dataframe_metabolite
# def signaling_meta_calculation(abun_path,upex_path,meta_path):


def filter_data_aundance_v6_OneDim(dataframe_abundance, dataframe_metadata, filter_params, comp_fact, comp_fact_value):
    if comp_fact=='ONE_DIMENSIONAL': #if it is ONE_DIMENSIONAL there is no need for filters
        filtered_df = dataframe_abundance
        return filtered_df
    if (not comp_fact==None) or (not comp_fact_value==None):
        filter_params[comp_fact] = comp_fact_value
    acceptable_filter_params = {}
    keys_filter_params = list(filter_params.keys())
    # print('filter_params',filter_params)
    # print('dataframe_metadata.columns.to_list()',dataframe_metadata.columns.to_list())


    for key in keys_filter_params:
        if not filter_params[key]=='' :
            if key in dataframe_metadata.columns.to_list():
                acceptable_filter_params[key] = filter_params[key]
    filtered_dataframe_metadata = dataframe_metadata.copy()
    for key in list(acceptable_filter_params.keys()):
        filtered_dataframe_metadata = filtered_dataframe_metadata[filtered_dataframe_metadata[key] == acceptable_filter_params[key]]
    allowed_sample_ids = filtered_dataframe_metadata.index.to_list()

    if not dataframe_abundance.empty:
        dataframe_abundance_all_sample_ids = dataframe_abundance.columns.to_list()
        Redundent_sample_ids = set(dataframe_abundance_all_sample_ids) - set(allowed_sample_ids)
        filtered_dataframe_abundance = dataframe_abundance.T.drop(list(Redundent_sample_ids)).T
        filtered_df = filtered_dataframe_abundance
    else:
        filtered_df = filtered_dataframe_metadata# this only filters meta-data so that when the signaling metabolite algorithm
                                                # is running, it automatically remove unrelevant values from comp_fact column

    return filtered_df




def iterative_zscore(dataframe_abundance1,dataframe_abundance2,dataframe_metabolite,G_metabolic,sample_number=15,iteration=10):

    Metabolite_Scores = {}
    for each_metabolite in dataframe_metabolite.index.to_list():
        Metabolite_Scores[each_metabolite] = []
    Metabolite_Scores_OneByOne =[]
    Reporter_Metabolites = {}

    #Select the lowest number of Samples when there is not enough sample
    if len(dataframe_abundance1.columns)<sample_number:
        print('There are not %i Samples in dataframe_abundance1'%(sample_number))
        sample_number = len(dataframe_abundance1.columns)
    if len(dataframe_abundance2.columns)<sample_number:
        print('There are not %i Samples in dataframe_abundance2'%(sample_number))
        sample_number = len(dataframe_abundance2.columns)
    print('Sample number is %i'%(sample_number))
    if sample_number<1: #for One-Dimension situation not happen
        print('Empty Data is not acceptable')
        return None, None

    #Iterative Z_Score Calculation
    for i_iter in range(iteration):
        print('iter %s==================================='%(str(i_iter)))
        #Random Samples from both abundances
        df_abundance1 = dataframe_abundance1.sample(n=sample_number,axis='columns')
        df_abundance2 = dataframe_abundance2.sample(n=sample_number,axis='columns')

        #Rename Sample names (to find difference)
        if sample_number != 1: #Checks if it should consider One_dim or Multi_dim Signaling_Metabolite
            diff_column_names = []
            for i in range(sample_number):
                diff_column_names.append('Sample%i'%(i))
            df_abundance1.columns = diff_column_names
            df_abundance2.columns = diff_column_names

            df_diff_data = abs(df_abundance1-df_abundance2)
        else: #sample_number == 1
            df_diff_data = df_abundance1
            df_diff_data.columns = ['p_values']


        for each_metabolite in dataframe_metabolite.index.to_list():
            if sample_number != 1:
                Z_metabolite_value, _ = Z_meta_Calculator_NOinf_v2(G_metabolic, each_metabolite, df_diff_data, Multi_dim=True)
            else: #sample_number == 1
                Z_metabolite_value, _ = Z_meta_Calculator_NOinf_v2(G_metabolic, each_metabolite, df_diff_data, Multi_dim=False)
            # print('Z_metabolite_value',Z_metabolite_value)
            #print(Z_metabolite_OneByOne)
            if not math.isnan(Z_metabolite_value):
                Metabolite_Scores[each_metabolite].append(Z_metabolite_value)
#                 Metabolite_Scores[each_metabolite] = (Z_metabolite_value)
#                 Metabolite_Scores_OneByOne.append(Z_metabolite_OneByOne)

    #Average of each iteration for Metabolite Score
    for each_metabolite in dataframe_metabolite.index.to_list():
        if np.isnan(np.average(Metabolite_Scores[each_metabolite])):
            del Metabolite_Scores[each_metabolite]
        else:
            Metabolite_Scores[each_metabolite] = np.average(Metabolite_Scores[each_metabolite])
    return Metabolite_Scores, Metabolite_Scores_OneByOne



def Z_meta_Calculator_NOinf_v2(G, metabolite, dataframe_abundance, Multi_dim=True,
                              Corr_method='spearman'):
    if Multi_dim == False :
#         for one dimentional situation
        Z_ej_all = []
        connected_species = list(G.neighbors(metabolite))
        for each_conncted_species in connected_species:
            Z_ej_all.append(dataframe_abundance['p_values'].loc[each_conncted_species])

        Z_ej_sum = [np.abs(stats.norm.ppf(np.double(1-x))) for x in Z_ej_all]
        # Z_ej_sum = np.array(Z_ej_sum)
        # Z_ej_sum = stats.norm.ppf(Z_ej_sum)

        Z_meta = np.sum(np.abs(Z_ej_sum))/np.sqrt(len(Z_ej_sum))
        Z_scores_OneByOne = [str(np.abs(np.double(x))) for x in Z_ej_all]
        return Z_meta, Z_scores_OneByOne
    else: #Multi_dim == True
#         for multi dimentional situation
        Z_ej_all = []
        connected_species = list(G.neighbors(metabolite))
        # print('connected_species',connected_species)
        for i_each_conncted_species in range(len(connected_species)):
            iname_each_conncted_species = connected_species[i_each_conncted_species]
            for j_each_conncted_species in range(i_each_conncted_species,len(connected_species)):
                jname_each_conncted_species = connected_species[j_each_conncted_species]
                if i_each_conncted_species==j_each_conncted_species: continue
                # try: #in case when a connected species is not available for this comp_fact_comb
                if Corr_method=='spearman':
                    # print('dataframe_abundance',dataframe_abundance)
                    # print('dataframe_abundance.iloc[j_each_conncted_species].to_list()',dataframe_abundance.iloc[j_each_conncted_species].to_list())
                    Pj = np.abs(spearmanr(dataframe_abundance.loc[iname_each_conncted_species].to_list(),
                                          dataframe_abundance.loc[jname_each_conncted_species].to_list())[0])
                elif Corr_method=='pearson':
                    Pj = np.abs(pearsonr(dataframe_abundance.loc[iname_each_conncted_species].to_list(),
                                         dataframe_abundance.loc[jname_each_conncted_species].to_list())[0])
                # except:
                #     continue
                # if No_inverse==False:
                #     Z_ej = stats.norm.ppf(Pj)
                # else:
                Z_ej = np.abs(Pj)
                Z_ej_all.append(Z_ej)
#                 print(Pj)
        Z_ej_sum = [np.abs(np.double(x)) for x in Z_ej_all]
        Z_ej_sum = np.array(Z_ej_sum)
        Z_ej_sum[np.where(Z_ej_sum == 0.0)[0]]=np.nan
        Z_ej_sum = stats.norm.ppf(Z_ej_sum)
        Z_ej_sum = np.ma.array(Z_ej_sum, mask=np.isnan(Z_ej_sum))

        Z_meta = np.sum(np.abs(Z_ej_sum))/np.sqrt(len(Z_ej_sum))
        Z_scores_OneByOne = [str(np.abs(np.double(x))) for x in Z_ej_all]
        return Z_meta, Z_scores_OneByOne




def Zscore_metabolite_correcter(Metabolite_Scores,k=25,iterations=20):
    print('correcting raw Z-scores...')
    metabolites = list(Metabolite_Scores.keys())
    mu_k_itrations = []
    sigma_k_itrations = []
    if k>len(metabolites):
        k = len(metabolites)

    for i in range(iterations):
        np.random.seed()
        selected_metabolites_index = np.random.permutation(len(metabolites))[:k]
        selected_scores = [Metabolite_Scores[metabolites[x]] for x in selected_metabolites_index]
        mu_k_itrations.append(np.mean(selected_scores))
        sigma_k_itrations.append(np.std(selected_scores))

    mu_k = np.mean(mu_k_itrations)
    sigma_k = np.std(sigma_k_itrations)

    Metabolite_Scores_Corrected = Metabolite_Scores.copy()
    for each_metabolite in metabolites:
        Metabolite_Scores_Corrected[each_metabolite] = (Metabolite_Scores_Corrected[each_metabolite] - mu_k)/sigma_k
    print('done!')
    return Metabolite_Scores_Corrected





    # Function : file_Exceler
def file_exceler_SigMeta(job_id,suffix_count=1):
    '''
    function : file_exceler_allinOneSheet
    args : inp_file_names : list of filenames to be zipped
    out_zip_file : output zip file adress
    return : none
    assumption : Input file paths and this code is in same directory.
    '''

    out_excel_file_address = 'Output_folder/'

    path = 'app_files/temp/'

    all_suffixes, all_files = find_all_sameNames(path,suffix_count)

    for each_suffix in all_suffixes:
        if each_suffix==job_id:
            inp_file_names = same_suffix(each_suffix,all_files,path)

    print(all_suffixes)
    try:

        #Create Excel file
        wb=Workbook()

        #This variable decides that how many times (for different Conditions) each Metabolite
        #should be more than Percentile_85 to be considered as Important Signaling Metabolite
        Number_Percentile_Thresh = 0 #

        #Create JSON file for Network Presentation
        JSON_net_2node = {'nodes':[],'links':[]}

        all_Zscores = np.array([])
        all_Zscores_df = pd.DataFrame()

        for file_to_write in inp_file_names:
            print(f' *** Processing file {file_to_write}')
            #load csv file
            csv_dataframe = pd.read_csv(file_to_write, header=[0], index_col=[0])

            #name of the sheet
            sheet_name = file_to_write.split('/')[-1] #just keep "suffix_count" last parts
            sheet_name = sheet_name.split('_')[suffix_count:] #just keep the file with .csv
            sheet_name[-1] = '.'.join(sheet_name[-1].split('.')[:-1])
            sheet_name = '_'.join(sheet_name)
            print(sheet_name)

            #Create a new sheet
            sheet = wb.create_sheet(sheet_name)

            #add Columns
            columns_all=[]
            columns_all.append(' ')
            [columns_all.append(x) for x in csv_dataframe.columns.to_list()]
            sheet.append(columns_all)

            #add other Rows
            for i_row in range(len(csv_dataframe)):
                each_row=[]
                each_row.append(csv_dataframe.index[i_row])
                [each_row.append(x) for x in csv_dataframe.iloc[i_row].to_list()]
                sheet.append(each_row)

            #Analyzing Signaling Metabolites
            zscore_csv = csv_dataframe
            #Column Name
            file_name = sheet_name
            #Separate Z_metabolite as vector
            zscore_csv_rawZscr = zscore_csv['Z_metabolite']
            #Append all Z_metabolites to find value of Percentile %85
            all_Zscores = np.append(all_Zscores,zscore_csv_rawZscr.to_numpy())
            #Append Z_metabolites for each file to a single DataFrame
            row = zscore_csv['Z_metabolite'].T
            row.name = file_name
            all_Zscores_df = all_Zscores_df.append(row)


        # Flip all_Zscores to Metabolites be the Indexes
        all_Zscores_df = all_Zscores_df.T
        # Find Percentile 85
        percentile_85_val = np.percentile(all_Zscores,q=85)
        # Find metabolites that at least have
        Important_Signaling_Metabolites = all_Zscores_df[(all_Zscores_df>=percentile_85_val).T.sum()>Number_Percentile_Thresh]
        #Creating Nodes and Links for JSON_net_3node


        #Adding Mean and STD
        Mean_col=Important_Signaling_Metabolites.mean(axis=1)
        Mean_col.name = 'Mean'
        STD_col=Important_Signaling_Metabolites.std(axis=1)
        STD_col.name = 'STD'
        Important_Signaling_Metabolites = Important_Signaling_Metabolites.T.append(Mean_col).T
        Important_Signaling_Metabolites = Important_Signaling_Metabolites.T.append(STD_col).T
        #Sort by Mean
        Important_Signaling_Metabolites = Important_Signaling_Metabolites.sort_values('Mean',ascending=False)
        #Adding nodes for JSON_net_3node
        for each_condition_node in Important_Signaling_Metabolites.columns.to_list():
            if each_condition_node=='Mean' or each_condition_node=='STD':
                continue
            JSON_net_2node['nodes'].append({"name":"%s"%(each_condition_node), "size":"%f"%(1.0), "color":"#0c0"})
        for each_metabolite_node in Important_Signaling_Metabolites.index.to_list():
            min_mean = Important_Signaling_Metabolites['Mean'].min(axis=0)
            max_mean = Important_Signaling_Metabolites['Mean'].max(axis=0)
            val = Important_Signaling_Metabolites['Mean'].loc[each_metabolite_node]
            node_size = 2*((val-min_mean)/(max_mean-min_mean))+1
            JSON_net_2node['nodes'].append({"name":"%s"%(each_metabolite_node), "size":"%f"%(node_size), "color":"#c00"})

        for source in Important_Signaling_Metabolites.columns.to_list():
            if source == 'Mean' or source == 'STD':
                continue
            for target in Important_Signaling_Metabolites.index.to_list():
                if Important_Signaling_Metabolites[source].loc[target] >= percentile_85_val:
                    JSON_net_2node['links'].append({"source":"%s"%(source),"target":"%s"%(target)})

        #Save JSON_net_2node
        path_json = os.path.join('Output_folder','Network_Effective','%s.json'%(job_id))
        with open(path_json, 'w') as fp:
            json.dump(JSON_net_2node, fp)

        # Get CVS file Ready for ParetoPlot
        pareto_ready(Important_Signaling_Metabolites,job_id)


        #Create a new sheet
        sheet = wb.create_sheet('Above_85_Percentile')

        #add Columns
        columns_all=[]
        columns_all.append(' ')
        [columns_all.append(x) for x in Important_Signaling_Metabolites.columns.to_list()]
        sheet.append(columns_all)

        #add other Rows
        for i_row in range(len(Important_Signaling_Metabolites)):
            each_row=[]
            each_row.append(Important_Signaling_Metabolites.index[i_row])
            [each_row.append(x) for x in Important_Signaling_Metabolites.iloc[i_row].to_list()]
            sheet.append(each_row)

        #Remove Default first sheet
        sheet = wb.get_sheet_by_name(wb.sheetnames[0])
        wb.remove(sheet)
    #         wb.remove_sheet()
        #save Excel file
        wb.save(os.path.join(out_excel_file_address,'%s.xlsx'%(job_id)))

        # removes all files
        for each_file in all_files:
            try:
                os.remove(os.path.join(path,each_file))
            except:
                print('File not found to delete: %s'%(each_file))
        print('Saved as Excel file!')
    except FileNotFoundError as e:
        print(f' *** Exception occurred during excel process - {e}')




def same_suffix(suffix,files,path):
    same_suffixes = []
    for each_file in files:
        if suffix in each_file:
            same_suffixes.append(os.path.join(path,'%s'%(each_file)))
    return same_suffixes



def find_all_sameNames(path,suffix_count=1):
    files = os.listdir(path)

    all_suffixes = []
    for each_file in files:
        try:
            suffix = each_file.split('/')[-1] #just keep "suffix_count" last parts
            suffix_temp = suffix.split('_')[:suffix_count]
            if suffix_count>1:
                suffix_final = '_'.join(suffix) #join "suffix_count" last parts as one string
            else:
                suffix_final = suffix_temp[0]
            all_suffixes.append(suffix_final) #store last parts in a list
        except:
            print('Error!')

    print(all_suffixes)
    all_suffixes = unique(all_suffixes)

    return all_suffixes, files




def pareto_ready(Important_Signaling_Metabolites,job_id):
    Mean_Sigmeta = Important_Signaling_Metabolites['Mean'].to_list()

    Pareto_df = pd.DataFrame(Mean_Sigmeta,columns=['Amount'],index=Important_Signaling_Metabolites.index.to_list())
    Pareto_df.index.name = 'Category'
    save_path = os.path.join('Output_folder','Paretoplot','%s.csv'%(job_id))
    Pareto_df.to_csv(save_path)#NEED TO CHANGE!!!!
