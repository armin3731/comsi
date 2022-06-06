import pandas as pd #Managing data tables, columns and rows
import numpy as np #library for numerical matrix and operators
import matplotlib.pyplot as plt #Showing graphs and plots
import os #Paths and operating system addresses
import openpyxl #library for working with excel file
from openpyxl import load_workbook,Workbook #loading Excel File
from openpyxl.utils import get_column_letter #Reading values of Excel data cells
import networkx as nx #Creating netwrok graphs
from matplotlib import cm #Color maps in graphs
from scipy.stats import spearmanr
import time
import io
import json
from mathematical_funcs import filter_data_aundance_v6_OneDim

#SPARCC imports
from analysis_methods import basis_corr
from io_methods import read_txt, write_txt

#Turn off warnings
import warnings
warnings.filterwarnings('ignore')





# Function : file_Exceler
def file_exceler(inp_file_names, out_excel_file_address):
    '''
    function : file_compress
    args : inp_file_names : list of filenames to be zipped
    out_zip_file : output zip file adress
    return : none
    assumption : Input file paths and this code is in same directory.
    '''
    try:

        #Create Excel file
        wb=Workbook()

        for file_to_write in inp_file_names:
            print(f' *** Processing file {file_to_write}')
            #load csv file
            csv_dataframe = pd.read_csv(file_to_write, header=[0], index_col=[0])

            #name of the sheet
            sheet_name = file_to_write.split('/')[-1] #just keep "suffix_count" last parts
            sheet_name = sheet_name.split('_')[:-suffix_count] #just keep the file without .csv
            sheet_name = '_'.join(sheet_name)
            print(sheet_name)

            #Create a new sheet
            sheet = wb.create_sheet(sheet_name)

            #add Columns
            columns_all=[]
            columns_all.append(' ')
            [columns_all.append(x) for x in csv_dataframe.columns.to_list()]
            sheet.append(columns_all)

            #add other Rows
            for i_row in range(len(csv_dataframe)):
                each_row=[]
                each_row.append(csv_dataframe.index[i_row])
                [each_row.append(x) for x in csv_dataframe.iloc[i_row].to_list()]
                sheet.append(each_row)


    except FileNotFoundError as e:
        print(f' *** Exception occurred during excel process - {e}')
    finally:
        sheet = wb.get_sheet_by_name(wb.sheetnames[0])
        wb.remove(sheet)
#         wb.remove_sheet()
        #save Excel file
        wb.save(out_excel_file_address)

        print('Done!')


suffix_count = 1

#Algorithm Parameters
alpha = 1 #Alpha const in the paper
Abundance_Type = 'avg_column' #The column name for average abunance dataframe

#Thresholds for Cummunity influence Rank
Thetas = [0.0,0.05,0.4]

TEMP_FOLDER = os.path.join('app_files','temp_influ')

# ==========================================================




def Bacteria_Influence_v1(job_id, dataframe_abundance, dataframe_metabolite, dataframe_metadata, filter_params, comp_fact,comp_fact_comb,Sample_number=50,Iteration_number=15):
    '''
    Finds Bacteria Influences using the chosen conditions and databases.

    Input(s):
    dataframe_abundance : Dataframe of Abundance database
    dataframe_metabolite : Dataframe of UPEX database
    dataframe_metadata : Dataframe of META-Data database
    filter_params : Filtering Parametes. this will filter samples as prefered
    comp_fact : the column in META-Data that comparison in taking on but it is only used for
                filtering function and for Bacteria influence it only have one input
    comp_fac_comb : a parameter for filtering function
    Sample_number : number of sample in each iteration for SPARCC
    Iteration_number : number of iterations for SPARCC

    Output(s):
         ---         : ---
    '''

    dataframe_abundance = filter_data_aundance_v6_OneDim(dataframe_abundance.copy(), dataframe_metadata.copy() ,
                                                               filter_params, comp_fact=None ,comp_fact_value=None)

# ---> Save Ref_Abundance to CSV
    dataframe_abundance.to_csv(os.path.join(TEMP_FOLDER,'Ref_Abundance_%s.csv'%(job_id)))
# ---> Save Ref_UPEX to CSV
    dataframe_metabolite.to_csv(os.path.join(TEMP_FOLDER,'Ref_UPEX_matrix_%s.csv'%(job_id)))


    #We only need MSPs which are shared with Metabolite data
    metabolite_MSPs = dataframe_metabolite.columns.to_list()
    abundance_MSPs = dataframe_abundance.index.to_list()
    # if abundance_MSPs > mehttp://localhost:8888/lab/workspaces/auto-9/tree/Metabolic_Network/Toy_model/toy_model_v30_DandruffProblem.ipynb#%3E%3E-Save-Ref_UPEX_matrix-to-CSVtabolite_MSPs:
    Redundant_MSPs = set(abundance_MSPs) - set(metabolite_MSPs)
    dataframe_abundance = dataframe_abundance.drop(Redundant_MSPs)
    # if metabolite_MSPs > abundance_MSPs :
    Redundant_MSPs = set(metabolite_MSPs) - set(abundance_MSPs)
    dataframe_metabolite = dataframe_metabolite.T
    dataframe_metabolite = dataframe_metabolite.drop(Redundant_MSPs)
    dataframe_metabolite = dataframe_metabolite.T
    # For Influence calculation, we need to have a average column for abundances
    dataframe_abundance_original = dataframe_abundance.copy()
    dict_abundance = dataframe_abundance.astype('float').mean(axis=1,skipna=False,numeric_only=True)
    dataframe_abundance = pd.DataFrame(dict_abundance,columns=['avg_column'])


    print('number of shared Abundance species',len(dataframe_abundance.index.to_list()))
    print('number of shared Metabolite species',len(dataframe_metabolite.columns.to_list()))

# ---> Save Abundance to CSV
    dataframe_abundance.to_csv(os.path.join(TEMP_FOLDER,'Abundance_MATRIX_%s.csv'%(job_id)))
# ---> Save UPEX to CSV
    dataframe_metabolite.to_csv(os.path.join(TEMP_FOLDER,'UPEX_MATRIX_%s.csv'%(job_id)))

    # Run Bacteria Influence
    output_influence_matrixes_v2(dataframe_abundance_org=dataframe_abundance, dataframe_metabolite=dataframe_metabolite,suffix=job_id,Thetas=Thetas)
    SPARCC_for_abundance(dataframe_abundance_original,job_id)

    files = os.listdir(TEMP_FOLDER)

    all_suffixes = []
    for each_file in files:
        try:
            suffix = each_file.split('_')[-suffix_count:] #just keep "suffix_count" last parts
            # suffix_temp = suffix[-1].split('.')[0]
            # suffix[-1] = suffix_temp
            suffix = '_'.join(suffix) #join "suffix_count" last parts as one string
            all_suffixes.append(suffix) #store last parts in a list
        except:
            print('Error! suffix seperate')


    all_suffixes = unique(all_suffixes)
    print('all_suffixes',all_suffixes)

    # for each_suffix in all_suffixes:
    try:
        input_file_list = same_suffix(job_id,files,TEMP_FOLDER)
        input_file_list = sort_for_Aras(input_file_list)
        output_file_address = os.path.join('Output_folder','%s.xlsx'%(job_id))
        file_exceler(input_file_list,output_file_address)
    #     print(outpu3t_file_address)
        for each_file in input_file_list:
            try:
                os.remove(each_file)
            except:
                print('File not found to delete: %s'%(each_file))
        print('Saved as Excel file!')
    except:
        print('Error! not Save as EXCEL or can Not Remove them')







def Abundace_finder(MSP_name, Abundance_Type, dataframe_abundance):
    '''
    Finds abundace for each type and msp

    INPUT
    ------------
        MSP_name        : MSP name
        Abundance_Type  : Abuandce type or one of the Headers of
                            dataframe_abundance ('Relative abundance (LGC)' or
                            'Relative abundance (HGC)' or 'Relative abundance (all)')

    OUTPUT
    ------------
        abundance       : Required abundance
    '''
    abundance = dataframe_abundance.loc[MSP_name][Abundance_Type]
    return abundance

def W_function_v202204(dataframe_abundance,Abundance_Type, dataframe_metabolite, Threshold=0.0):
    '''
    Calculates direct influence of each metabolite on others one by one
    Upex matrix can have 3 valid values : 1  ,  -1   ,  1/-1 ,   5
      1 : when a species produces a metabolite
     -1 : when a species consumes a metabolite
    1/-1: when a species both produces and consumes a metabolite
      5 : When a species degrade a macro molecule

    INPUT
    ------------
        Abundance_Type  : Abuandce type or one of SampleIDs that direct influence is
                          calculating on
        dataframe       : the dataframe which contains used metabolites for each MSP
        Threshold       : Wij = 0   if    (Wij <= Threshold)

    OUTPUT
    ------------
        W               : Direct influence Matrix
        W_graph         : Direct influence for each two MSPs in Graph format
        dataframe_metabolite_abundances : normalized values for each metabolite
        W_df            : Direct influnece Matrix in Dataframe format (for better presentation)
        direct_metaboliete_weight_df : W_by_Metabolite which is consist of coefficients for uptake and release entities
    '''
    dataframe_metabolite_abundances = dataframe_metabolite.copy() # A copy of dataframe_metabolite in order to reserve
                                                                  # Relative Abundaces for echa MSP in every Metabolite
    dataframe_metabolite = dataframe_metabolite.astype('string')

    MSP = dataframe_abundance.index.to_list() #Name of existing MSPs
    MSP_number = len(MSP) #Total number of MSPs

    metabolites = dataframe_metabolite.index.to_list()#Name of existing Metabolites

    W = np.zeros([MSP_number,MSP_number])   #Initial W_ij matrix with zero values
    W_graph = [] #Initial W_graph to vizualize graph
    W_graph_positive = []
    W_graph_negative = []

    temp_single_row_abundance = np.zeros([1,MSP_number],dtype=float) #This is an empty row that can reserve
                                                                     #Relative Abundances for each row

    for k in range(len(dataframe_metabolite)): #for every metabolite in database
        temp_total_uptake = 0.  #an empty variable to reserve total abundance of uptake entities
        temp_total_release = 0. #an empty variable to reserve total abundance of release entities
    #STEP1 ---------Calculation------------------------------------------------------------
        #A loop for calculating Total Abundances of
        #Uptake and Release metabolites for each row
        #NOTE: -1 is Uptake and 1 is Release
        for each_MSP in MSP:

            #uptake_release_cell_value is the value of each cell
            #to determin if it is -1 or 1 (uptake or release)
            #for each MSP
            uptake_release_cell_value = dataframe_metabolite[each_MSP].iloc[k]


            if str(uptake_release_cell_value).strip() == "1": #Uptake
                temp_total_uptake += float(Abundace_finder(each_MSP,Abundance_Type, dataframe_abundance))
            elif str(uptake_release_cell_value).strip() == "-1": #Release
                temp_total_release += float(Abundace_finder(each_MSP,Abundance_Type,dataframe_abundance))
#                 print("-1 --- Metabolite: %s ,  MSP: %s"%(dataframe_metabolite.index.to_list()[k],each_MSP))
            elif str(uptake_release_cell_value).strip() == "1/-1": #Both
                temp_total_uptake += float(Abundace_finder(each_MSP,Abundance_Type, dataframe_abundance))
                temp_total_release += float(Abundace_finder(each_MSP,Abundance_Type,dataframe_abundance))
            else: #if uptake_release_cell_value is 0
                continue # jumps to next MSP


    #end_STEP1-------------------------------------------------------------------


    #STEP2 ----------------------------------------------------------------------
        #A loop to calculate Relative Abundance for
        #for each cell in a row
        for each_MSP in MSP:

            uptake_release_cell_value = dataframe_metabolite[each_MSP].iloc[k] #Value of each cell
#             dataframe_metabolite_abundances[each_MSP].iloc[k] = dataframe_metabolite_abundances[each_MSP].iloc[k]
            #Relative Abundances for Uptake and Release entities
            if str(uptake_release_cell_value).strip() == "1": #Uptake
                #if temp_total_uptake or temp_total_release = 0 then it means that for
                #this SAMPLE, there was no Abundance. Therefore, there is no weight for
                #this metabolite.
                if not temp_total_uptake == 0.0:
                    #alpha * (n_i / SIGMA{n_m})
                    dataframe_metabolite_abundances[each_MSP].iloc[k] = (((alpha * float(Abundace_finder(each_MSP,Abundance_Type,dataframe_abundance)))/temp_total_uptake))
                else:
                    dataframe_metabolite_abundances[each_MSP].iloc[k] = 0.0

            elif str(uptake_release_cell_value).strip() == "-1": #Release
                #if temp_total_uptake or temp_total_release = 0 then it means that for
                #this SAMPLE, there was no Abundance. Therefore, there is no weight for
                #this metabolite.
                if not temp_total_release == 0.0:
                    #-1 * (n_i / SIGMA{n_m})
                    dataframe_metabolite_abundances[each_MSP].iloc[k] = (((-1 * float(Abundace_finder(each_MSP,Abundance_Type,dataframe_abundance)))/temp_total_release))
                else:
                    dataframe_metabolite_abundances[each_MSP].iloc[k] = 0.0
            elif str(uptake_release_cell_value).strip() == "1/-1": #Release
                #if the species both produce and consumes (1/-1) then based on formula it both
                #weights should calculate and subtract from each other
                temp_total_weight = 0.
                if not temp_total_uptake == 0.0:
                    #alpha * (n_i / SIGMA{n_m})
                    temp_total_weight += (((alpha * float(Abundace_finder(each_MSP,Abundance_Type,dataframe_abundance)))/temp_total_uptake))
                else:
                    temp_total_weight += 0.0
                if not temp_total_release == 0.0:
                    #-1 * (n_i / SIGMA{n_m})
                    temp_total_weight += (((-1 * float(Abundace_finder(each_MSP,Abundance_Type,dataframe_abundance)))/temp_total_release))
                else:
                    temp_total_weight += 0.0
                dataframe_metabolite_abundances[each_MSP].iloc[k] = temp_total_weight
            elif str(uptake_release_cell_value).strip() == "5": #Macro Molecule
                #Nothing has decided for Macro Molecules right now
                dataframe_metabolite_abundances[each_MSP].iloc[k] = 0.0
            else: #if uptake_release_cell_value is 0
                continue # jumps to next MSP

    w_by_metabolite_input = dataframe_metabolite_abundances.copy()
    w_by_metabolite_input = w_by_metabolite_input.astype('float')
    #end_STEP2-------------------------------------------------------------------






    #in order to calculate W_ij, two loops are inplemented for j
    for j in range(MSP_number):   # j MSP
        j_name = MSP[j] #Name of j MSP


    #STEP3 ----------------------------------------------------------------------
        #Filter data: when j==-1 and i!=0
        #         dataframe_metabolite[j_name] = pd.to_numeric(dataframe_metabolite[j_name],downcast='float')
        dataframe_metabolite_filtered = dataframe_metabolite[((dataframe_metabolite[j_name])=="-1") | (dataframe_metabolite[j_name]=="1/-1") ] #filte data where j is -1

        #Metabolites that species j is consuming
        Consumed_metaolites_j = dataframe_metabolite_filtered.index.to_list()
    #end_STEP3-------------------------------------------------------------------


    #STEP4 ----------------------------------------------------------------------
        #filtered data (based on consumed metabolites by species j)
        # are summed up
        for i in range(MSP_number):    # i MSP
            i_name = MSP[i] #Name of i MSP
            proposed_Wij = np.sum(dataframe_metabolite_abundances[i_name].loc[Consumed_metaolites_j].to_numpy().astype('float')) # W_ij
            if (i==j)or(np.abs(proposed_Wij)<=Threshold): continue #jumps to next MSP if i=j
            W[i,j] = proposed_Wij
            #W_graph for i and j to use in networkx library
            W_graph.append([j_name,i_name,str(W[i,j])])
    #end_STEP4 ------------------------------------------------------------------

    #W as dataframe for saving it much easier as CSV file
    W_df = pd.DataFrame(W, columns=MSP, index=MSP)
#     print(W)

    #STEP-ALPHA------------------------------------------------------------------
    #Calculate W_by_metabolite

    #Bacteria names and combinations of their names two by two
    Bacteries = MSP#Every bacteria available
    Bacterie_combined_names = []

    for each_bacteria1 in Bacteries:
        for each_bacteria2 in Bacteries:
            if each_bacteria1==each_bacteria2:continue
            Bacterie_combined_names.append(each_bacteria1+'|'+each_bacteria2)#combine every bacteria two by two

    #Set empty matrixes in order to calculate W_by_metabolites
    direct_metaboliete_weight_matrix = np.zeros([len(metabolites), (len(Bacteries)*len(Bacteries))-len(Bacteries)])#Empty matrix for w_by_metabolite data
    direct_metaboliete_weight_df = pd.DataFrame(direct_metaboliete_weight_matrix, columns=Bacterie_combined_names, index=metabolites)#empty dataframe for w_by_metabolite data

    for each_bacteria1 in Bacteries:#source bacterie
        for each_bacteria2 in Bacteries:#target bacterie
            if each_bacteria1==each_bacteria2:continue
            Consumed_metabolites_by_each_Bac = w_by_metabolite_input[(w_by_metabolite_input[each_bacteria1]<0) & ~(w_by_metabolite_input[each_bacteria2]==0)] #filter data to find in which metabolites j is consummer and i is not 0
            try:
                involved_metabolites = Consumed_metabolites_by_each_Bac.index.to_list()#Metabolites that are consumed by each_bacteria1(source) and has a weight with each_bacteria2(target)
                for each_involved_metabolite in involved_metabolites:
                    direct_metaboliete_weight_df[each_bacteria1+'|'+each_bacteria2].loc[each_involved_metabolite] = Consumed_metabolites_by_each_Bac[each_bacteria2].loc[each_involved_metabolite]
            except:
                #there is no shared metabolites
                continue
    #end_STEP-ALPHA------------------------------------------------------------------

    return W, W_graph, dataframe_metabolite_abundances, W_df, direct_metaboliete_weight_df




#Defining the graph
def w2graph(W_graph, name='MSP Graph'):
    G=nx.DiGraph(name=name)
    interactions = np.array(W_graph)
    for i in range(len(interactions)): #adding nodes and weights to the graph
        interaction = interactions[i]
        a = interaction[0] # First MSP (source node)
        b = interaction[1] # Second MSP (target node)
        ########################ABS VALUES#########################!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        w = float(interaction[2]) # weight
        G.add_weighted_edges_from([(a,b,w)]) # add weighted edge to graph
    return G


# function to get unique values
def unique(list1):
    unique_vals = []
    # insert the list to the set
    list_set = set(list1)
    # convert the set to the list
    unique_list = (list(list_set))
    for x in unique_list:
        unique_vals.append(x)
    return unique_vals


# Find the shortest path between two MSPs and calculate the total effect
#number_of_shortestpass: 0 when there is a direct connection
#                         -1 when there is no connection
#                        OTHER NUMBERS whern there are multiple bacterias in between
def shortestpath_effect_v4(graph, source, target, deltaNiNi=1., theta_phi=0., connection_details=False):
    try:
        shortest_paths = [p for p in nx.all_shortest_paths(graph, source=source, target=target)]
        if connection_details:print(shortest_paths)
        Avg_weight = []
        if (len(shortest_paths)==1)and(len(shortest_paths[0])<=2):
            if connection_details:print('There is a direct Connection')
            return 0., 1, 0,'DC'
        shortpath_rank = len(shortest_paths[0])-1
        number_of_shortestpath = len(shortest_paths)
        for each_path in shortest_paths:
            path_weight = 1
            for each_node_idx in range(len(each_path)-1):
                path_weight *= graph.edges[(each_path[each_node_idx], each_path[each_node_idx+1])]['weight']
            path_weight = np.abs(path_weight)############################################################
            Avg_weight.append(path_weight)
        Total_effect = np.average(Avg_weight)
        Total_effect *= deltaNiNi
        Total_effect -= theta_phi
        return Total_effect, shortpath_rank, number_of_shortestpath, shortest_paths
    except nx.NetworkXNoPath:
        if connection_details:print('There is no Connection between source and target nodes')
        return 0. , 0, 0, 'NC'
    except nx.NodeNotFound:
        if connection_details:print('%s dose not have any connection at all'%(source))
        return 0. , 0, 0, 'NC_atALL'


def deltaNi_calculator(source_msp, Normal_column_name, Disease_column_name):
    ni_normal = float(Abundace_finder(source_msp, Normal_column_name,dataframe_abundance))
    ni_disease = float(Abundace_finder(source_msp, Disease_column_name,dataframe_abundance))

    deltaNi = np.abs(ni_normal-ni_disease)
    deltaNiNi = deltaNi/ni_normal
    return deltaNiNi


def output_influence_matrixes_v2(dataframe_abundance_org,dataframe_metabolite,suffix='job_id',Thetas=[0.0,0.05]):


    dataframe_abundance = dataframe_abundance_org.copy() ##Backup the not filtered data (GEOGRAPHY filter)


    #START finding the outputs---------------------
    #Checks if all MSPs in Abundance data exist in Metabolite data
    if not len(dataframe_abundance.index.to_list()) == len(dataframe_metabolite.columns.to_list()):
        if len(dataframe_abundance.index.to_list()) > len(dataframe_metabolite.columns.to_list()):
            redundant_MSPs = set(dataframe_abundance.index.to_list()) - set(dataframe_metabolite.columns.to_list())
            dataframe_abundance = dataframe_abundance.drop(redundant_MSPs)
        else:
            redundant_MSPs = set(dataframe_metabolite.columns.to_list()) - set(dataframe_abundance.index.to_list())
            dataframe_metabolite = dataframe_metabolite.T.drop(redundant_MSPs).T

    #SAVE: Abundance_HMA_SW_Control to CSV
#     dataframe_abundance.to_csv('output/matrix/Abundance_HMA_%s_%s_%s.csv'%(Geo,Case_Control,Disease))

    #SAVE: UPEX_matrix_SW_Control to CSV
#     dataframe_metabolite.to_csv('output/matrix/UPEX_matrix_%s_%s_%s.csv'%(Geo,Case_Control,Disease))


    Bacteries = dataframe_abundance.index.to_list()#All Bacteries

    #Empty variables=================================
    indexes_DIM = []
    columns_DIM = []
    indexes_InDIM = []
    columns_InDIM = []
    indexes_CIN = []
    columns_CIN = []

    W_bibj = np.zeros([len(Bacteries),len(Bacteries)])
    W_bibj_ind = np.zeros([len(Bacteries),len(Bacteries)])
    Shortpass_dist = np.zeros([len(Bacteries),len(Bacteries)])
    Shortpass_names = [['' for i in range(len(Bacteries))] for j in range(len(Bacteries))]
    number_of_bacteria_effected_indirect = np.zeros([len(Bacteries),1])
    number_of_bacteria_effected_total = np.zeros([len(Bacteries),1+2])
    Community_positive_values = np.zeros([len(Bacteries),len(Bacteries)])
    Community_negative_values = np.zeros([len(Bacteries),len(Bacteries)])
    number_of_bacteria_effected_total = np.zeros([len(Bacteries),len(Thetas)])
    #end Empty variables=============================


    #Direct and InDirect Influences ===========================================
    each_sample = dataframe_abundance.columns[0]
    W_s1,Wgraph_s1,data_for_check,W_s1_df, W_by_metabolites = W_function_v202204(dataframe_abundance,each_sample,dataframe_metabolite)
    G1 = w2graph(Wgraph_s1)
    d3js_graph_maker(Wgraph_s1, all_nodes=Bacteries ,job_id=suffix)#save the graph

    for i in range(len(Bacteries)):
        for j in range(len(Bacteries)):
            if Bacteries[i]==Bacteries[j]:continue
    #Direct Influence
            W_bibj[i,j] = W_s1_df[Bacteries[j]].loc[Bacteries[i]]

    #Indirect Influence
            source_msp = Bacteries[i]
            target_msp = Bacteries[j]
            indirect_effect , shortpass_distance_, _, shortpass_name_ = shortestpath_effect_v4(G1,source=source_msp,target=target_msp,connection_details=False)#, deltaNiNi=deltaNiNi)
            W_bibj_ind[i,j] = indirect_effect

    #Shortest Paths:
    ##Shortpass Distance
            Shortpass_dist[i,j] = shortpass_distance_
            Shortpass_names[i][j] = shortpass_name_

    ##Community Positive and Negative values
            if W_bibj[i,j]+W_bibj_ind[i,j] >= 0:
                Community_positive_values[i,j] = W_bibj[i,j]+W_bibj_ind[i,j]
            else:
                Community_negative_values[i,j] = W_bibj[i,j]+W_bibj_ind[i,j]

    ##Community Rank
            for i_theta in range(len(Thetas)):
                each_theta = Thetas[i_theta]
                number_of_bacteria_effected_total[i,i_theta] += np.heaviside(np.abs(W_bibj_ind[i,j])+np.abs(W_bibj[i,j])-each_theta,0.0)

    #Indexes
    for i in range(len(Bacteries)):
        indexes_DIM.append(Bacteries[i])
        indexes_InDIM.append(Bacteries[i])
        indexes_CIN.append(Bacteries[i])

    #Columns
        columns_DIM.append(Bacteries[i])
        columns_InDIM.append(Bacteries[i])
        columns_CIN.append(Bacteries[i])

    Direct_Influence_Matrix_df = pd.DataFrame(W_bibj, columns=columns_DIM, index=indexes_DIM)
    Indirect_Influence_Matrix_df = pd.DataFrame(W_bibj_ind, columns=columns_InDIM, index=indexes_InDIM)

    #SAVE: Direct_Influence to CSV
    Direct_Influence_Matrix_df.to_csv(os.path.join(TEMP_FOLDER,'_Dir_Influ_MATIRX_%s.csv'%(suffix)))
    print('Direct Influence Done!')

    #SAVE: Indirect_Influence to CSV
    Indirect_Influence_Matrix_df.to_csv(os.path.join(TEMP_FOLDER,'_InDir_Influ_MATIRX_%s.csv'%(suffix)))
    print('InDirect Influence Done!')

    #Total Influence
    Total_influence_weight = Indirect_Influence_Matrix_df + Direct_Influence_Matrix_df

    #SAVE: Total_Influence to CSV
    Total_influence_weight.to_csv(os.path.join(TEMP_FOLDER,'Total_Influence_MATIRX_%s.csv'%(suffix)))
    print('Total Influence Done!')

    Shortpass_Distance = pd.DataFrame(Shortpass_dist, columns=columns_CIN, index=indexes_CIN)
    #SAVE: Shortpath_Distant to CSV
    Shortpass_Distance.to_csv(os.path.join(TEMP_FOLDER,'Shortpath_Distant_MATIRX_%s.csv'%(suffix)))
    print("Shortest Path Distance Done!")

    Shortpass_MSPs = pd.DataFrame(Shortpass_names, columns=columns_CIN, index=indexes_CIN)
    #SAVE: Shortpath_MSPs to CSV
    Shortpass_MSPs.to_csv(os.path.join(TEMP_FOLDER,'Shortpath_MSPs_MATIRX_%s.csv'%(suffix)))
    print("Shortest Path MSPs Done!")

    #SAVE: Direct Influence by Metabolite to CSV
    W_by_metabolites.to_csv(os.path.join(TEMP_FOLDER,'Dir_Influ_by_Metabolite_%s.csv'%(suffix)))
    print('Direct Influence by Metabolite Done!')

    for i_theta in range(len(Thetas)):
        each_Theta = Thetas[i_theta]
        Community_Influence_number_df = pd.DataFrame(np.squeeze(number_of_bacteria_effected_total[:,i_theta]), index=Bacteries)
        #SAVE: Community_Rank to CSV
        Community_Influence_number_df.to_csv(os.path.join(TEMP_FOLDER,'Community_Rank_Thresh%0.2f_%s.csv'%(i_theta,suffix)))
    print("Community Influence Number Done!")

    bacteria_effected_total_positive_df = pd.DataFrame(Community_positive_values, columns=Bacteries, index=Bacteries)
    #SAVE: Community_Influence_Weight_Positive to CSV
    bacteria_effected_total_positive_df.to_csv(os.path.join(TEMP_FOLDER,'Community_Weight_Positive_MATIRX_%s.csv'%(suffix)))
    print("Positive Weights Done!")

    bacteria_effected_total_negative_df = pd.DataFrame(Community_negative_values, columns=Bacteries, index=Bacteries)
    #SAVE: Community_Influence_Weight_Negative to CSV
    bacteria_effected_total_negative_df.to_csv(os.path.join(TEMP_FOLDER,'Community_Weight_Negative_MATIRX_%s.csv'%(suffix)))
    print("Negative Weight Done!")

#     data_for_check.to_csv('output/matrix/data_for_check.csv')
    return 'Done!'


    # Function : file_Exceler
def file_exceler(inp_file_names, out_excel_file_address):
    '''
    function : file_compress
    args : inp_file_names : list of filenames to be zipped
    out_zip_file : output zip file adress
    return : none
    assumption : Input file paths and this code is in same directory.
    '''
    try:
        print(inp_file_names)
        #Create Excel file
        wb=Workbook()

        for file_to_write in inp_file_names:
            print(f' *** Processing file {file_to_write}')
            #load csv file
            csv_dataframe = pd.read_csv(file_to_write, header=[0], index_col=[0])

            #name of the sheet
            sheet_name = file_to_write.split('/')[-1] #just keep "suffix_count" last parts
            sheet_name = sheet_name.split('_')[:-suffix_count] #just keep the file without .csv
            sheet_name = '_'.join(sheet_name)
            print(sheet_name)

            #Create a new sheet
            sheet = wb.create_sheet(sheet_name)

            #add Columns
            columns_all=[]
            columns_all.append(' ')
            [columns_all.append(x) for x in csv_dataframe.columns.to_list()]
            sheet.append(columns_all)

            #add other Rows
            for i_row in range(len(csv_dataframe)):
                each_row=[]
                each_row.append(csv_dataframe.index[i_row])
                [each_row.append(x) for x in csv_dataframe.iloc[i_row].to_list()]
                sheet.append(each_row)


    except FileNotFoundError as e:
        print(f' *** Exception occurred during excel process - {e}')
    finally:
        sheet = wb.get_sheet_by_name(wb.sheetnames[0])
        wb.remove(sheet)
#         wb.remove_sheet()
        #save Excel file
        wb.save(out_excel_file_address)

        print('Done!')




def same_suffix(suffix,files,path):
    same_suffixes = []
    for each_file in files:
        if suffix in each_file:
            same_suffixes.append(os.path.join(path,'%s'%(each_file)))
    return same_suffixes



def sort_for_Aras(input_file_list):
    sort_order = ['Ref_Abundance','Ref_UPEX','Abundance_MATRIX','UPEX_MATRIX','_Dir_Influ_','Shortpath_Distant_','Shortpath_MSPs_','_InDir_Influ_','Total_Influence_','Dir_Influ_by_Metabolite','Community_Rank','Community_Weight_','Centrality_','COR_SparCC','COV_SparCC','Metabolite_interactions']
    re_arrange_names = []
    for each_sort_order_member in sort_order:
        while (each_sort_order_member in '|'.join(input_file_list)):
            for each_input_file_list in input_file_list:
                if each_sort_order_member in each_input_file_list:
                    re_arrange_names.append(each_input_file_list)
                    input_file_list.remove(each_input_file_list)

    return re_arrange_names



def SPARCC_for_abundance(dataframe_abundance,job_id):
    sample_number = len(dataframe_abundance.columns.to_list())

    if sample_number > 1 :
        dataframe_abundance_transpose = dataframe_abundance.T
        print('computing correlations')
        cor, cov = basis_corr(dataframe_abundance_transpose, method='SparCC', iter=25)#, **kwargs)
        cor.to_csv(os.path.join(TEMP_FOLDER,'COR_SparCC_%s.csv'%(job_id)))
        cov.to_csv(os.path.join(TEMP_FOLDER,'COV_SparCC_%s.csv'%(job_id)))
        print('SPARCC Done!')


def d3js_graph_maker(Wgraph, all_nodes ,job_id):
    #Species NODES:
    JSON_net_2node = {'nodes':[],'links':[]}
    for each_node in all_nodes:
        JSON_net_2node['nodes'].append({"name":"%s"%(each_node), "size":"%f"%(1.0), "color":"#00c"})

    #Add Edges======================
    for each_link in Wgraph:
            JSON_net_2node['links'].append({"source":"%s"%(each_link[0]),"target":"%s"%(each_link[1]),"weight":"%s"%(each_link[2])})

    #Save JSON_net_2node
    path_json = os.path.join('Output_folder','Network_Normal','%s.json'%(job_id))
    with open(path_json, 'w') as fp:
        json.dump(JSON_net_2node, fp)
